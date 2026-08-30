"""
Crowdfunding Tracker — suivi des investissements en crowdfunding immobilier
(Bricks.co, La Première Brique, ...)

FastAPI + SQLite. Single user (auth par cookie de session).
Données saisies manuellement (sync automatique = futur bonus).
"""

import csv
import hashlib
import io
import json
import os
import re
import secrets
import sqlite3
import unicodedata
import urllib.request
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Optional

from fastapi import FastAPI, HTTPException, Request, Response
from fastapi.responses import FileResponse, JSONResponse, Response
from fastapi.staticfiles import StaticFiles
from pydantic import BaseModel, Field

# ---------------------------------------------------------------- config

BASE_DIR = Path(__file__).resolve().parent.parent
DATA_DIR = Path(os.environ.get("DATA_DIR", BASE_DIR / "data"))
DB_PATH = Path(os.environ.get("DB_PATH", DATA_DIR / "app.db"))
STATIC_DIR = BASE_DIR / "public"

PORT = int(os.environ.get("PORT", "8016"))

ADMIN_USER = os.environ.get("ADMIN_USER", "admin")
ADMIN_PASSWORD = os.environ.get("ADMIN_PASSWORD", "change-me")

SESSION_TTL_DAYS = 30

PLATFORMS = {
    "bricks": "Bricks.co",
    "lapremierebrique": "La Première Brique",
}

# Statuts possibles
STATUS_LABELS = {
    "en_collecte": "En collecte",
    "en_cours": "En cours",
    "retard": "⚠️ Retard de paiement",
    "rembourse": "Remboursé",
    "perdu": "Perte / Défaut",
}

app = FastAPI(title="Crowdfunding Tracker")

# ---------------------------------------------------------------- db helpers


def db() -> sqlite3.Connection:
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    return conn


def init_db() -> None:
    DATA_DIR.mkdir(parents=True, exist_ok=True)
    conn = db()
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS users (
            username TEXT PRIMARY KEY,
            pwd_hash TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS sessions (
            token TEXT PRIMARY KEY,
            username TEXT NOT NULL,
            expires_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS projects (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT NOT NULL DEFAULT 'bricks',
            name TEXT NOT NULL,
            city TEXT DEFAULT '',
            invested REAL NOT NULL DEFAULT 0,
            rate REAL NOT NULL DEFAULT 0,
            duration_months INTEGER NOT NULL DEFAULT 0,
            start_date TEXT,
            expected_end_date TEXT,
            actual_end_date TEXT,
            status TEXT NOT NULL DEFAULT 'en_cours',
            repaid_capital REAL NOT NULL DEFAULT 0,
            interest_received REAL NOT NULL DEFAULT 0,
            reinvested_from INTEGER,
            notes TEXT DEFAULT '',
            auto_created INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        );
        CREATE TABLE IF NOT EXISTS operations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            platform TEXT NOT NULL,
            source_id TEXT,
            op_date TEXT NOT NULL,
            type TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'Validée',
            project_id INTEGER,
            amount REAL NOT NULL DEFAULT 0,
            details TEXT DEFAULT '',
            contract_type TEXT DEFAULT '',
            extra TEXT DEFAULT '',
            created_at TEXT NOT NULL,
            UNIQUE(platform, source_id)
        );
        CREATE INDEX IF NOT EXISTS idx_ops_project ON operations(project_id);
        CREATE INDEX IF NOT EXISTS idx_ops_date ON operations(op_date);
        """
    )
    # seed admin user
    row = conn.execute("SELECT 1 FROM users WHERE username=?", (ADMIN_USER,)).fetchone()
    if row is None:
        conn.execute(
            "INSERT INTO users (username, pwd_hash) VALUES (?, ?)",
            (ADMIN_USER, hash_password(ADMIN_PASSWORD)),
        )
    # migrations idempotentes (CREATE TABLE IF NOT EXISTS n'ajoute pas les colonnes)
    try:
        conn.execute("ALTER TABLE projects ADD COLUMN auto_created INTEGER NOT NULL DEFAULT 0")
    except sqlite3.OperationalError:
        pass  # colonne déjà présente
    conn.execute(
        """CREATE TABLE IF NOT EXISTS sync_tokens (
            id INTEGER PRIMARY KEY CHECK (id=1),
            token TEXT NOT NULL,
            created_at TEXT NOT NULL
        )"""
    )
    conn.execute(
        """CREATE TABLE IF NOT EXISTS scrape_captures (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            ts TEXT NOT NULL,
            platform TEXT DEFAULT '',
            url TEXT DEFAULT '',
            status_code INTEGER DEFAULT 0,
            body TEXT DEFAULT ''
        )"""
    )
    conn.execute(
        """CREATE TABLE IF NOT EXISTS scrape_report (
            id INTEGER PRIMARY KEY CHECK (id=1),
            data TEXT DEFAULT '{}',
            created_at TEXT NOT NULL
        )"""
    )
    # colonnes additionnelles (scraping) — ALTER idempotent
    for col in (
        "interest_net REAL DEFAULT 0",
        "interest_remaining REAL DEFAULT 0",
        "interest_remaining_net REAL DEFAULT 0",
        "real_rate REAL DEFAULT 0",
        "contract_type TEXT DEFAULT ''",
        "valuation REAL DEFAULT 0",
        "rest_months INTEGER DEFAULT 0",
        "infine INTEGER DEFAULT 0",
    ):
        try:
            conn.execute(f"ALTER TABLE projects ADD COLUMN {col}")
        except Exception:
            pass
    # métadonnées par plateforme : solde dispo, total déposé, valeur dans les projets
    conn.execute("""CREATE TABLE IF NOT EXISTS platform_meta (
        platform TEXT PRIMARY KEY,
        balance REAL DEFAULT 0,
        deposited REAL DEFAULT 0,
        invested_value REAL DEFAULT 0,
        updated_at TEXT DEFAULT ''
    )""")
    # valeurs initiales (éditables dans l'UI) : totaux déposés communiqués par l'utilisateur
    conn.execute("INSERT OR IGNORE INTO platform_meta (platform, balance, deposited, invested_value) VALUES ('bricks', 0, 2390, 3628.88)")
    conn.execute("INSERT OR IGNORE INTO platform_meta (platform, balance, deposited, invested_value) VALUES ('lapremierebrique', 0, 1500, 0)")
    # indices de comparaison (taux annualisés, période 09/2021 → 08/2026)
    conn.execute("""CREATE TABLE IF NOT EXISTS benchmarks (
        key TEXT PRIMARY KEY,
        name TEXT NOT NULL,
        annual_pct REAL NOT NULL,
        note TEXT DEFAULT ''
    )""")
    for k, name, val, note in [
        ("sp500", "S&P 500", 11.86, ""),
        ("nasdaq", "Nasdaq Composite", 12.22, ""),
        ("msci_world", "MSCI World", 11.86, "ETF iShares (IWDA)"),
        ("stoxx600", "STOXX Europe 600", 7.34, ""),
        ("cac40", "CAC 40", 4.86, ""),
        ("livret_a", "Livret A", 2.20, "taux réglementé moyen"),
    ]:
        conn.execute("INSERT OR IGNORE INTO benchmarks (key, name, annual_pct, note) VALUES (?,?,?,?)",
                     (k, name, val, note))
    # simulation « ETF Acc » : perf cumulée réelle de chaque indice depuis la date du
    # 1er achat de chaque plateforme (Bricks 2021-09-27, LPB 2022-10-03) jusqu'au
    # dernier cours dispo (2026-08-28) — prix Yahoo Finance, Livret A = taux composé
    for col in ("cum_bricks REAL DEFAULT 0", "cum_lpb REAL DEFAULT 0", "asof TEXT DEFAULT ''"):
        try:
            conn.execute(f"ALTER TABLE benchmarks ADD COLUMN {col}")
        except Exception:
            pass
    for k, cb, cl in [
        ("sp500", 0.7357, 1.0965),
        ("nasdaq", 0.7637, 1.4412),
        ("msci_world", 0.7354, 1.1803),
        ("stoxx600", 0.4168, 0.6763),
        ("cac40", 0.2632, 0.4499),
        ("livret_a", 0.1132, 0.0888),
    ]:
        conn.execute("UPDATE benchmarks SET cum_bricks=?, cum_lpb=?, asof='2026-08-28' WHERE key=?",
                     (cb, cl, k))
    # historique mensuel des perfs cumulées par indice (pour la courbe « Évolution du
    # patrimoine ») — source : prix Yahoo Finance, fichier src/index_history.json
    conn.execute("CREATE TABLE IF NOT EXISTS index_history (key TEXT NOT NULL, month TEXT NOT NULL, cum_bricks REAL, cum_lpb REAL, PRIMARY KEY (key, month))")
    ih_file = BASE_DIR / "src" / "index_history.json"
    if ih_file.exists():
        try:
            ih_data = json.loads(ih_file.read_text())
            for k, months in ih_data.items():
                for m, (cb, cl) in months.items():
                    conn.execute("INSERT OR REPLACE INTO index_history (key, month, cum_bricks, cum_lpb) VALUES (?,?,?,?)",
                                 (k, m, cb, cl))
            conn.commit()
        except Exception:
            pass
    conn.commit()
    conn.close()


def hash_password(pwd: str, salt: Optional[str] = None) -> str:
    if salt is None:
        salt = secrets.token_hex(16)
    digest = hashlib.pbkdf2_hmac("sha256", pwd.encode(), salt.encode(), 200_000).hex()
    return f"{salt}${digest}"


def verify_password(pwd: str, stored: str) -> bool:
    try:
        salt, digest = stored.split("$", 1)
    except ValueError:
        return False
    return secrets.compare_digest(hash_password(pwd, salt), stored)


def now_iso() -> str:
    return datetime.utcnow().isoformat(timespec="seconds") + "Z"


# ---------------------------------------------------------------- auth

def make_session(response: Response, username: str) -> None:
    token = secrets.token_hex(32)
    expires = (datetime.utcnow() + timedelta(days=SESSION_TTL_DAYS)).isoformat()
    conn = db()
    conn.execute(
        "INSERT INTO sessions (token, username, expires_at) VALUES (?, ?, ?)",
        (token, username, expires),
    )
    conn.commit()
    conn.close()
    response.set_cookie(
        "ct_session",
        token,
        max_age=SESSION_TTL_DAYS * 86400,
        httponly=True,
        samesite="lax",
        secure=os.environ.get("COOKIE_SECURE", "0") == "1",
    )


def current_user(request: Request) -> Optional[str]:
    token = request.cookies.get("ct_session")
    if not token:
        return None
    conn = db()
    row = conn.execute(
        "SELECT username FROM sessions WHERE token=? AND expires_at > ?",
        (token, datetime.utcnow().isoformat()),
    ).fetchone()
    conn.close()
    return row["username"] if row else None


def require_user(request: Request) -> str:
    user = current_user(request)
    if not user:
        raise HTTPException(status_code=401, detail="Non authentifié")
    return user


# ---------------------------------------------------------------- business logic

def add_months(d: date, months: int) -> date:
    m = d.month - 1 + months
    y = d.year + m // 12
    m = m % 12 + 1
    day = min(d.day, [31, 29 if y % 4 == 0 and (y % 100 != 0 or y % 400 == 0) else 28,
                      31, 30, 31, 30, 31, 31, 30, 31, 30, 31][m - 1])
    return date(y, m, day)


def parse_date(s: Optional[str]) -> Optional[date]:
    if not s:
        return None
    try:
        return date.fromisoformat(s)
    except ValueError:
        return None


def sync_indicators_from_ops(conn) -> dict:
    """Recalcule interest_received et le statut depuis les opérations importées
    (source de vérité = exports) pour les projets auto-créés.
    - interest_received = somme des revenus reçus (types « Revenus », hors revente)
    - statut 'rembourse' si une opération de revente totale / remboursement final existe
    """
    stats = {"interest": 0, "status": 0}
    try:
        projects = conn.execute("SELECT id, name FROM projects WHERE auto_created=1").fetchall()
        for p in projects:
            row = conn.execute(
                """SELECT COALESCE(SUM(amount),0) tot FROM operations
                   WHERE project_id=? AND status IN ('Validée','Réussi')
                     AND type LIKE '%Revenus%' AND type NOT LIKE '%revente%' AND amount > 0""",
                (p["id"],)).fetchone()
            if row and row["tot"]:
                conn.execute("UPDATE projects SET interest_received=? WHERE id=?",
                             (round(row["tot"], 2), p["id"]))
                stats["interest"] += 1
            # projet terminé : revente totale ou remboursement final du capital
            done = conn.execute(
                """SELECT 1 FROM operations WHERE project_id=?
                   AND status IN ('Validée','Réussi')
                   AND (type LIKE '%revente totale%' OR type LIKE '%Remboursement final%' OR type LIKE '%remboursement final%')""",
                (p["id"],)).fetchone()
            if done:
                conn.execute("UPDATE projects SET status='rembourse' WHERE id=? AND status='en_cours'",
                             (p["id"],))
                stats["status"] += 1
            # invested = souscriptions − annulations de souscription (LPB) :
            # une souscription annulée (« Réussi » +X €) n'est PAS un investissement
            row = conn.execute(
                """SELECT
                     COALESCE(SUM(CASE WHEN amount<0 AND type LIKE '%Souscription%'
                                      THEN -amount ELSE 0 END),0) subs,
                     COALESCE(SUM(CASE WHEN amount>0 AND type LIKE '%nnulation%'
                                      THEN amount ELSE 0 END),0) annul
                   FROM operations WHERE project_id=? AND status IN ('Validée','Réussi')""",
                (p["id"],)).fetchone()
            if row and (row["subs"] or row["annul"]):
                inv = round(row["subs"] - row["annul"], 2)
                conn.execute("UPDATE projects SET invested=? WHERE id=?",
                             (inv, p["id"]))
        conn.commit()
    except Exception:
        pass
    return stats


def project_computed(row: sqlite3.Row, today: Optional[date] = None, extra: Optional[dict] = None) -> dict:
    """Calcule tous les indicateurs dérivés d'un projet."""
    today = today or date.today()
    p = dict(row)
    extra = extra or {}

    start = parse_date(p.get("start_date"))
    expected_real = parse_date(p.get("expected_end_date"))
    expected = expected_real
    # échéance dérivée (uniquement pour l'affichage, JAMAIS pour le retard) :
    # 1) « X mois max. restants » LPB (info officielle de la plateforme)
    # 2) sinon start + durée du contrat
    derived = None
    if expected is None and p.get("rest_months"):
        derived = add_months(today, int(p["rest_months"]))
        expected = derived
    if expected is None and start is not None and p.get("duration_months"):
        derived = add_months(start, int(p["duration_months"]))
        expected = derived

    invested = p.get("invested") or 0.0
    rate = p.get("rate") or 0.0
    repaid = p.get("repaid_capital") or 0.0
    interest_recv = p.get("interest_received") or 0.0
    status = p.get("status") or "en_cours"

    # --- retard (auto) : SEULEMENT si l'échéance est réelle (scrapée/saisie),
    # pas dérivée de start+durée (évite les faux retards)
    is_late = False
    days_delayed = 0
    if status in ("en_cours",) and expected_real is not None:
        days_delayed = (today - expected_real).days
        is_late = days_delayed > 0

    # --- intérêts courus (latents, non encore versés)
    accrued = 0.0
    if status == "en_cours" and start is not None and invested > 0 and rate > 0:
        days = (today - start).days
        if days > 0:
            accrued = round(invested * (rate / 100.0) * days / 365.0, 2)

    # --- gravité du retard (projets « retard » uniquement) : intérêts REÇUS vs
    # intérêts ATTENDUS depuis le début (capital × taux × durée écoulée)
    expected_interest = None
    interest_ratio = None
    late_severity = None
    if status == "retard":
        exp = None
        if start is not None and invested > 0 and rate > 0:
            d = (today - start).days
            if d > 0:
                exp = round(invested * (rate / 100.0) * d / 365.0, 2)
        expected_interest = exp
        if exp:
            interest_ratio = round(interest_recv / exp, 3)
        if p.get("infine"):
            late_severity = "negligeable"  # in-fine : intérêts versés au terme (normal)
        elif exp is not None:
            ratio = interest_ratio if interest_ratio is not None else 0.0
            if interest_recv <= 0 or ratio < 0.3:
                late_severity = "critique"      # aucun / quasi aucun intérêt perçu
            elif ratio < 0.6:
                late_severity = "important"
            elif ratio < 0.9:
                late_severity = "significatif"
            else:
                late_severity = "negligeable"   # les paiements suivent à peu près
        else:
            # pas de taux/début connus : on juge sur la fraîcheur des paiements
            last_int = extra.get("last_interest")
            if last_int and (today - last_int).days < 120:
                late_severity = "significatif"  # paie mais statut « retard »
            else:
                late_severity = "important"
    months_since_interest = None
    if extra.get("last_interest"):
        months_since_interest = round((today - extra["last_interest"]).days / 30.44, 1)

    # --- perte (statut perdu)
    loss = 0.0
    if status == "perdu":
        loss = round(max(0.0, invested - repaid), 2)

    # --- perte latente : mise investie > valeur actuelle des bricks
    # (mark-to-market, projets actifs uniquement)
    valuation = p.get("valuation") or 0
    unrealized_loss = 0.0
    if valuation > 0 and invested > valuation and status in ("en_cours", "retard", "en_collecte"):
        unrealized_loss = round(invested - valuation, 2)

    capital_due = round(max(0.0, invested - repaid), 2) if status in ("en_cours", "en_collecte") else 0.0

    # --- taux réel annualisé (projets terminés uniquement) : tout ce qui est
    # revenu (capital + intérêts, source de vérité = opérations) vs mise,
    # annualisé sur la durée RÉELLE (début → fin réelle)
    total_received = None
    real_annual_pct = None
    if status in ("rembourse", "perdu"):
        recv_map = extra.get("total_received") or {}
        recv = recv_map.get(p.get("id"))
        if recv:
            total_received = round(float(recv), 2)
        else:
            total_received = round(repaid + interest_recv, 2)
        fin = p.get("actual_end_date") or (extra.get("last_op_date") or {}).get(p.get("id")) or p.get("expected_end_date")
        fin_d = parse_date(fin)
        if invested > 0 and start is not None and fin_d is not None:
            j = (fin_d - start).days
            if j > 0:
                ratio = total_received / invested - 1.0
                real_annual_pct = -1.0 if ratio <= -1.0 else (1.0 + ratio) ** (365.0 / j) - 1.0

    total_gain = round(interest_recv + accrued, 2)
    net = round(total_gain - loss, 2)

    return {
        **p,
        "expected_end_date": expected_real.isoformat() if expected_real else None,
        "derived_end_date": derived.isoformat() if derived else None,
        "is_late": is_late,
        "days_delayed": max(0, days_delayed),
        "accrued_interest": accrued,
        "expected_interest": expected_interest,
        "interest_ratio": interest_ratio,
        "late_severity": late_severity,
        "months_since_interest": months_since_interest,
        "unrealized_loss": unrealized_loss,
        "loss": loss,
        "capital_due": capital_due,
        "total_gain": total_gain,
        "net": net,
        "total_received": total_received,
        "real_annual_pct": real_annual_pct,
        "platform_label": PLATFORMS.get(p.get("platform"), p.get("platform")),
    }


# ---------------------------------------------------------------- import (xlsx)

# Types Bricks.co qui représentent un investissement (mise)
BRICKS_INVEST_TYPES = {"Achat de bricks", "Achat marketplace", "Frais d'achat marketplace"}
# Types LPB qui représentent une souscription (mise)
LPB_INVEST_PREFIX = "Souscription au projet"
LPB_CANCEL_PREFIX = "Annulation de la souscription au projet"
LPB_REPAY_PREFIX = "Remboursement mensualité"


def _norm_name(s: Optional[str]) -> str:
    """Normalise un nom de projet : minuscules, sans accents, apostrophes unifiées."""
    if not s:
        return ""
    import unicodedata
    s = unicodedata.normalize("NFD", str(s))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().replace("’", "'").replace(" ", "").strip()


def _parse_fr_amount(v) -> float:
    """Parse un montant '35,56 €' / '-100,00 €' / 35.56 / -10.18."""
    if v is None:
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("€", "").replace("\u00a0", "").replace(" ", "").strip()
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_fr_date(v) -> Optional[str]:
    """Parse '10/08/2026' → '2026-08-10'. Retourne None si non parsable."""
    if v is None:
        return None
    s = str(v).strip()
    for fmt in ("%d/%m/%Y", "%d/%m/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s[:10], fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _norm_apos(s: str) -> str:
    """Normalise les apostrophes typographiques (’ U+2019) en apostrophe droite."""
    return str(s).replace("\u2019", "'").replace("\u2018", "'")


def _parse_bricks_rows(rows) -> list:
    """Rows Bricks.co : id, date, type, statut, propriété, type de contrat, montant (€), prix de la brick (€)."""
    ops = []
    for r in rows:
        if not r or not r[0]:
            continue
        source_id = str(r[0]).strip()
        op_date = _parse_fr_date(r[1])
        if not op_date:
            continue
        ops.append({
            "source_id": source_id,
            "op_date": op_date,
            "type": _norm_apos(r[2] or "").strip(),
            "status": str(r[3] or "Validée").strip(),
            "project_name": _norm_apos(r[4] or "").strip() or None,
            "contract_type": str(r[5] or "").strip(),
            "amount": _parse_fr_amount(r[6]),
            "extra": {"brick_price": r[7]},
        })
    return ops


def _parse_lpb_rows(rows) -> list:
    """Rows LPB : Nature de la transaction, Moyen de paiement, Détails, Montant, Statut, Date d'exécution."""
    import re as _re
    ops = []
    for r in rows:
        if not r or not r[0]:
            continue
        nature = str(r[0] or "").strip()
        op_date = _parse_fr_date(r[5])
        if not op_date:
            continue
        m = _re.search(r"projet\s+(.+?)\s*$", nature)
        pname = m.group(1).strip() if m else None
        ops.append({
            "source_id": None,  # fingerprint calculé plus bas (pas d'id natif)
            "op_date": op_date,
            "type": nature,
            "status": str(r[4] or "Réussi").strip(),
            "project_name": pname,
            "contract_type": str(r[1] or "").strip(),
            "amount": _parse_fr_amount(r[3]),
            "extra": {},
        })
    # fingerprint : hash(date|nature|montant|statut) pour idempotence
    for o in ops:
        fp = hashlib.sha1(
            f"{o['op_date']}|{o['type']}|{o['amount']}|{o['status']}".encode()
        ).hexdigest()
        o["source_id"] = fp
    return ops


def _detect_platform(headers) -> Optional[str]:
    h = [str(x or "").strip().lower() for x in headers]
    if any("prix de la brick" in x for x in h) or "id" in h and "propriété" in h:
        return "bricks"
    if any("nature de la transaction" in x for x in h):
        return "lapremierebrique"
    return None


def _find_or_create_project(conn, platform: str, pname: str, invested: float, start_date: str) -> int:
    """Retrouve un projet par nom normalisé (même plateforme), sinon le crée (auto_created)."""
    key = _norm_name(pname)
    row = conn.execute(
        "SELECT id, auto_created FROM projects WHERE platform=? AND name=?",
        (platform, pname.strip()),
    ).fetchone()
    if row is None:
        # matching normalisé en dernier recours
        rows = conn.execute("SELECT id, name, auto_created FROM projects WHERE platform=?", (platform,)).fetchall()
        for r in rows:
            if _norm_name(r["name"]) == key:
                row = r
                break
    if row is not None:
        pid = row["id"]
        # projet auto-créé → on met à jour mise/date si l'import apporte plus d'infos
        if row["auto_created"]:
            cur = conn.execute("SELECT invested, start_date FROM projects WHERE id=?", (pid,)).fetchone()
            new_invested = round(cur["invested"] + invested, 2)
            new_start = start_date
            if cur["start_date"] and cur["start_date"] < start_date:
                new_start = cur["start_date"]
            conn.execute(
                "UPDATE projects SET invested=?, start_date=?, updated_at=? WHERE id=?",
                (new_invested, new_start, now_iso(), pid),
            )
        return pid
    cur = conn.execute(
        """INSERT INTO projects
           (platform, name, city, invested, rate, duration_months, start_date,
            expected_end_date, actual_end_date, status, repaid_capital,
            interest_received, reinvested_from, notes, auto_created, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (platform, pname.strip(), "", invested, 0, 0, start_date, None, None,
         "en_cours", 0, 0, None, "Importé automatiquement depuis l'export de la plateforme",
         1, now_iso(), now_iso()),
    )
    return cur.lastrowid


def _import_operations(conn, platform: str, ops: list) -> dict:
    """Insère les opérations + crée/maj les projets. Retourne un résumé.

    Les projets ne sont créés/mis à jour qu'à partir des opérations d'investissement
    RÉELLEMENT nouvelles (source_id absent de la DB) — un ré-import du même fichier
    ne doit jamais re-majorer la mise d'un projet.
    """
    summary = {"imported": 0, "duplicates": 0, "projects_created": 0, "projects_updated": 0, "warnings": []}

    def _is_invest(o: dict) -> bool:
        if o["status"] not in ("Validée", "Réussi"):
            return False  # opérations annulées/refusées/échouées : l'argent n'a jamais bougé
        if platform == "bricks":
            return o["type"] in BRICKS_INVEST_TYPES
        return o["type"].startswith(LPB_INVEST_PREFIX)

    # source_ids déjà présents (idempotence)
    existing = {r["source_id"] for r in conn.execute(
        "SELECT source_id FROM operations WHERE platform=?", (platform,)
    ).fetchall()}

    # 1) investissements NOUVEAUX par projet → création/mise à jour des projets
    inv_new: dict[str, dict] = {}
    for o in ops:
        if not o["project_name"] or not _is_invest(o):
            continue
        if o["source_id"] in existing:
            continue
        d = inv_new.setdefault(o["project_name"], {"total": 0.0, "first": o["op_date"]})
        d["total"] += abs(o["amount"])
        if o["op_date"] < d["first"]:
            d["first"] = o["op_date"]

    for pname, d in inv_new.items():
        before = conn.execute("SELECT COUNT(*) c FROM projects WHERE platform=?", (platform,)).fetchone()["c"]
        _find_or_create_project(conn, platform, pname, d["total"], d["first"])
        after = conn.execute("SELECT COUNT(*) c FROM projects WHERE platform=?", (platform,)).fetchone()["c"]
        if after > before:
            summary["projects_created"] += 1
        else:
            summary["projects_updated"] += 1

    # 2) opérations (INSERT OR IGNORE pour l'idempotence)
    for o in ops:
        pid = None
        if o["project_name"]:
            row = conn.execute(
                "SELECT id FROM projects WHERE platform=? AND name=?",
                (platform, o["project_name"].strip()),
            ).fetchone()
            if row is None:
                rows = conn.execute(
                    "SELECT id, name FROM projects WHERE platform=?", (platform,)
                ).fetchall()
                for r in rows:
                    if _norm_name(r["name"]) == _norm_name(o["project_name"]):
                        row = r
                        break
            pid = row["id"] if row else None
        cur = conn.execute(
            """INSERT OR IGNORE INTO operations
               (platform, source_id, op_date, type, status, project_id, amount,
                details, contract_type, extra, created_at)
               VALUES (?,?,?,?,?,?,?,?,?,?,?)""",
            (platform, o["source_id"], o["op_date"], o["type"], o["status"], pid,
             o["amount"], o["type"], o["contract_type"],
             __import__("json").dumps(o["extra"], ensure_ascii=False), now_iso()),
        )
        if cur.rowcount:
            summary["imported"] += 1
        else:
            summary["duplicates"] += 1

    conn.commit()
    return summary


# ---------------------------------------------------------------- models

class LoginIn(BaseModel):
    username: str
    password: str


class PasswordIn(BaseModel):
    old_password: str
    new_password: str = Field(min_length=4)


class ProjectIn(BaseModel):
    platform: str = "bricks"
    name: str = Field(min_length=1)
    city: str = ""
    invested: float = 0
    rate: float = 0
    duration_months: int = 0
    start_date: Optional[str] = None
    expected_end_date: Optional[str] = None
    actual_end_date: Optional[str] = None
    status: str = "en_cours"
    repaid_capital: float = 0
    interest_received: float = 0
    interest_net: float = 0
    interest_remaining: float = 0
    interest_remaining_net: float = 0
    real_rate: float = 0
    contract_type: str = ""
    valuation: float = 0
    reinvested_from: Optional[int] = None
    notes: str = ""


# ---------------------------------------------------------------- auth routes

@app.post("/api/auth/login")
def login(body: LoginIn, response: Response):
    conn = db()
    row = conn.execute("SELECT pwd_hash FROM users WHERE username=?", (body.username,)).fetchone()
    conn.close()
    if row is None or not verify_password(body.password, row["pwd_hash"]):
        raise HTTPException(status_code=401, detail="Identifiants invalides")
    make_session(response, body.username)
    return {"ok": True, "username": body.username}


@app.post("/api/auth/logout")
def logout(request: Request, response: Response):
    token = request.cookies.get("ct_session")
    if token:
        conn = db()
        conn.execute("DELETE FROM sessions WHERE token=?", (token,))
        conn.commit()
        conn.close()
    response.delete_cookie("ct_session")
    return {"ok": True}


@app.get("/api/auth/me")
def me(request: Request):
    user = current_user(request)
    return {"authenticated": user is not None, "username": user}


@app.post("/api/auth/change-password")
def change_password(body: PasswordIn, request: Request):
    user = require_user(request)
    conn = db()
    row = conn.execute("SELECT pwd_hash FROM users WHERE username=?", (user,)).fetchone()
    if row is None or not verify_password(body.old_password, row["pwd_hash"]):
        conn.close()
        raise HTTPException(status_code=400, detail="Ancien mot de passe incorrect")
    conn.execute("UPDATE users SET pwd_hash=? WHERE username=?", (hash_password(body.new_password), user))
    conn.commit()
    conn.close()
    return {"ok": True}


# ---------------------------------------------------------------- overview (performance globale)

@app.get("/api/overview")
def get_overview(request: Request):
    require_user(request)
    conn = db()
    conn.row_factory = sqlite3.Row
    today = date.today()
    metas = {r["platform"]: dict(r) for r in conn.execute("SELECT * FROM platform_meta")}
    # valeur dans les projets : LPB = capital restant dû (investi − remboursé)
    cap_due = {}
    for r in conn.execute("SELECT platform, SUM(MAX(0, invested - COALESCE(repaid_capital,0))) s FROM projects GROUP BY platform"):
        cap_due[r["platform"]] = r["s"] or 0
    # début des investissements (1er achat) → annualisation
    first_buy = {}
    for r in conn.execute("""SELECT p.platform, MIN(o.op_date) d FROM operations o JOIN projects p ON p.id=o.project_id
        WHERE o.amount<0 AND (o.type LIKE '%Achat%' OR o.type LIKE '%ouscription%') AND o.status IN ('Validée','Réussi')
        GROUP BY p.platform"""):
        first_buy[r["platform"]] = r["d"]
    out = []
    for plat in ("bricks", "lapremierebrique"):
        m = metas.get(plat, {})
        balance = round(m.get("balance") or 0, 2)
        deposited = round(m.get("deposited") or 0, 2)
        if plat == "lapremierebrique":
            invested_value = round(cap_due.get(plat) or 0, 2)  # capital restant dû (auto)
        else:
            invested_value = round(m.get("invested_value") or 0, 2)  # valeur des bricks (site/saisie)
        patrimoine = round(balance + invested_value, 2)
        gain = round(patrimoine - deposited, 2) if deposited else None
        ratio = round(gain / deposited, 4) if (deposited and gain is not None) else None
        annual = None
        d0 = first_buy.get(plat)
        if ratio is not None and d0:
            try:
                days = (today - date.fromisoformat(d0)).days
                if days > 0 and (1 + ratio) > 0:
                    annual = round(((1 + ratio) ** (365.0 / days) - 1) * 100, 2)
            except ValueError:
                pass
        out.append({
            "platform": plat,
            "balance": balance,
            "deposited": deposited,
            "invested_value": invested_value,
            "invested_value_auto": plat == "lapremierebrique",
            "patrimoine": patrimoine,
            "gain": gain,
            "ratio": ratio,
            "annual_pct": annual,
            "start_date": d0,
            "updated_at": m.get("updated_at") or "",
        })
    # total toutes plateformes (annualisé depuis le 1er achat global)
    t_dep = round(sum(p["deposited"] for p in out), 2)
    t_pat = round(sum(p["patrimoine"] for p in out), 2)
    t_gain = round(t_pat - t_dep, 2) if t_dep else None
    t_ratio = round(t_gain / t_dep, 4) if (t_dep and t_gain is not None) else None
    t_annual = None
    d0s = [p["start_date"] for p in out if p["start_date"]]
    if t_ratio is not None and d0s:
        try:
            days = (today - date.fromisoformat(min(d0s))).days
            if days > 0 and (1 + t_ratio) > 0:
                t_annual = round(((1 + t_ratio) ** (365.0 / days) - 1) * 100, 2)
        except ValueError:
            pass
    conn.close()
    # indices de comparaison
    conn2 = db()
    conn2.row_factory = sqlite3.Row
    benchmarks = [dict(r) for r in conn2.execute("SELECT * FROM benchmarks ORDER BY annual_pct DESC")]
    conn2.close()
    # simulation « ETF Acc » : les dépôts réels de chaque plateforme, placés à la
    # date de son 1er achat dans un ETF capitalisant suivant l'indice
    sim_dep = {p["platform"]: p["deposited"] for p in out}
    sim_tot_dep = round(sum(sim_dep.values()), 2)
    sim_asof = ""
    for b in benchmarks:
        val = 0.0
        for plat, dep in sim_dep.items():
            cum = b.get("cum_bricks") if plat == "bricks" else b.get("cum_lpb")
            val += dep * (1.0 + (cum or 0))
        b["sim_value"] = round(val, 2)
        b["sim_gain"] = round(val - sim_tot_dep, 2) if sim_tot_dep else None
        if b.get("asof"):
            sim_asof = b["asof"]
    return {"platforms": out, "benchmarks": benchmarks, "sim_asof": sim_asof, "total": {
        "deposited": t_dep, "patrimoine": t_pat, "gain": t_gain,
        "ratio": t_ratio, "annual_pct": t_annual,
    }}


@app.put("/api/platform-meta")
async def update_platform_meta(request: Request):
    require_user(request)
    b = await request.json()
    plat = (b.get("platform") or "").strip()
    if plat not in ("bricks", "lapremierebrique"):
        raise HTTPException(status_code=400, detail="Plateforme invalide")
    conn = db()
    conn.row_factory = sqlite3.Row
    cur = conn.execute("SELECT * FROM platform_meta WHERE platform=?", (plat,)).fetchone()
    vals = {k: (b.get(k) if b.get(k) is not None else (cur[k] if cur else 0)) for k in ("balance", "deposited", "invested_value")}
    conn.execute(
        "INSERT INTO platform_meta (platform, balance, deposited, invested_value, updated_at)"
        " VALUES (?,?,?,?,?) ON CONFLICT(platform) DO UPDATE SET balance=?, deposited=?, invested_value=?, updated_at=?",
        (plat, vals["balance"], vals["deposited"], vals["invested_value"], now_iso(),
         vals["balance"], vals["deposited"], vals["invested_value"], now_iso()))
    conn.commit()
    conn.close()
    return {"ok": True}


@app.put("/api/benchmarks")
async def update_benchmark(request: Request):
    require_user(request)
    b = await request.json()
    key = (b.get("key") or "").strip()
    val = b.get("annual_pct")
    if not key or val is None:
        raise HTTPException(status_code=400, detail="Clé et taux requis")
    conn = db()
    n = conn.execute("UPDATE benchmarks SET annual_pct=? WHERE key=?", (float(val), key)).rowcount
    conn.commit()
    conn.close()
    if n == 0:
        raise HTTPException(status_code=404, detail="Indice introuvable")
    return {"ok": True}


@app.post("/api/benchmarks")
async def add_benchmark(request: Request):
    """Ajoute un indice de comparaison ; perf cumulée (ETF Acc) estimée depuis le
    taux annualisé avec les mêmes dates de départ que la simulation."""
    require_user(request)
    b = await request.json()
    name = (b.get("name") or "").strip()
    val = b.get("annual_pct")
    if not name or val is None:
        raise HTTPException(status_code=400, detail="Nom et taux requis")
    try:
        val = float(val)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Taux invalide")
    conn = db()
    base = unicodedata.normalize("NFKD", name.lower()).encode("ascii", "ignore").decode()
    base = re.sub(r"[^a-z0-9_]+", "_", base).strip("_") or "indice"
    key, n = base, 2
    while conn.execute("SELECT 1 FROM benchmarks WHERE key=?", (key,)).fetchone():
        key, n = f"{base}_{n}", n + 1
    today = date.today()
    j1 = (today - date(2021, 9, 27)).days
    j2 = (today - date(2022, 10, 3)).days
    cb = round((1 + val / 100) ** (j1 / 365) - 1, 4)
    cl = round((1 + val / 100) ** (j2 / 365) - 1, 4)
    note = (b.get("note") or "").strip()
    conn.execute(
        "INSERT INTO benchmarks (key, name, annual_pct, note, cum_bricks, cum_lpb, asof) VALUES (?,?,?,?,?,?,?)",
        (key, name, val, note, cb, cl, today.isoformat()))
    conn.commit()
    conn.close()
    return {"ok": True, "key": key}


@app.delete("/api/benchmarks/{key}")
async def delete_benchmark(key: str, request: Request):
    require_user(request)
    conn = db()
    n = conn.execute("DELETE FROM benchmarks WHERE key=?", (key,)).rowcount
    conn.commit()
    conn.close()
    if n == 0:
        raise HTTPException(status_code=404, detail="Indice introuvable")
    return {"ok": True}


# ---------------------------------------------------------------- extension download

@app.get("/api/extension-zip")
def extension_zip(request: Request, browser: str = "chromium"):
    require_user(request)
    ext_dir = BASE_DIR / "extension"
    if not ext_dir.is_dir():
        raise HTTPException(status_code=404, detail="Extension non disponible")
    import io, json as _json, zipfile
    buf = io.BytesIO()
    with zipfile.ZipFile(buf, "w", zipfile.ZIP_DEFLATED) as z:
        for f in sorted(ext_dir.rglob("*")):
            if f.is_file():
                z.write(f, f.relative_to(ext_dir))
        if browser == "firefox":
            # manifest adapté Firefox (MV3 supporté depuis FF 109 ; id requis)
            m = _json.loads((ext_dir / "manifest.json").read_text(encoding="utf-8"))
            m["browser_specific_settings"] = {
                "gecko": {"id": "crowdfunding-tracker-sync@local", "strict_min_version": "115.0"}
            }
            z.writestr("manifest.json", _json.dumps(m, indent=2, ensure_ascii=False))
    buf.seek(0)
    fname = f"crowdfunding-tracker-sync-{browser}.zip"
    return Response(content=buf.getvalue(), media_type="application/zip",
                    headers={"Content-Disposition": f'attachment; filename="{fname}"'})


# ---------------------------------------------------------------- backup / restore

@app.get("/api/backup")
def backup(request: Request):
    require_user(request)
    conn = db()
    conn.row_factory = sqlite3.Row
    data = {
        "app": "crowdfunding-tracker",
        "version": (BASE_DIR / "VERSION").read_text().strip() if (BASE_DIR / "VERSION").exists() else "0.0.0",
        "exported_at": now_iso(),
        "projects": [dict(r) for r in conn.execute("SELECT * FROM projects ORDER BY id")],
        "operations": [dict(r) for r in conn.execute("SELECT * FROM operations ORDER BY id")],
        "platform_meta": [dict(r) for r in conn.execute("SELECT * FROM platform_meta")],
        "benchmarks": [dict(r) for r in conn.execute("SELECT * FROM benchmarks")],
    }
    conn.close()
    return data


@app.post("/api/restore")
async def restore(request: Request):
    require_user(request)
    body = await request.json()
    if not isinstance(body, dict) or "projects" not in body or "operations" not in body:
        raise HTTPException(status_code=400, detail="Sauvegarde invalide")
    conn = db()
    try:
        conn.execute("BEGIN")
        conn.execute("DELETE FROM operations")
        conn.execute("DELETE FROM projects")
        conn.execute("DELETE FROM platform_meta")
        conn.execute("DELETE FROM benchmarks")
        for r in body.get("platform_meta") or []:
            conn.execute(
                "INSERT OR REPLACE INTO platform_meta (platform, balance, deposited, invested_value, updated_at) VALUES (?,?,?,?,?)",
                (r.get("platform", ""), r.get("balance") or 0, r.get("deposited") or 0,
                 r.get("invested_value") or 0, r.get("updated_at") or now_iso()))
        for r in body.get("benchmarks") or []:
            conn.execute(
                "INSERT OR REPLACE INTO benchmarks (key, name, annual_pct, note) VALUES (?,?,?,?)",
                (r.get("key", ""), r.get("name", ""), r.get("annual_pct") or 0, r.get("note") or ""))
        for r in body["projects"]:
            conn.execute(
                """INSERT OR REPLACE INTO projects
                   (id, platform, name, city, invested, rate, duration_months, start_date,
                    expected_end_date, actual_end_date, status, repaid_capital,
                    interest_received, reinvested_from, notes, auto_created, created_at, updated_at,
                    real_rate, contract_type, valuation, rest_months, infine)
                   VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
                (r.get("id"), r.get("platform") or "bricks", r.get("name", ""), r.get("city") or "",
                 r.get("invested") or 0, r.get("rate") or 0, r.get("duration_months") or 0,
                 r.get("start_date"), r.get("expected_end_date"), r.get("actual_end_date"),
                 r.get("status") or "en_cours", r.get("repaid_capital") or 0,
                 r.get("interest_received") or 0, r.get("reinvested_from"),
                 r.get("notes") or "", r.get("auto_created") or 0,
                 r.get("created_at") or now_iso(), r.get("updated_at") or now_iso(),
                 r.get("real_rate") or 0, r.get("contract_type") or "",
                 r.get("valuation") or 0, r.get("rest_months") or 0, r.get("infine") or 0))
        for r in body["operations"]:
            conn.execute(
                """INSERT OR REPLACE INTO operations
                   (id, platform, project_id, type, op_date, amount, status, source_id, created_at)
                   VALUES (?,?,?,?,?,?,?,?,?)""",
                (r.get("id"), r.get("platform") or "bricks", r.get("project_id"),
                 r.get("type", ""), r.get("op_date"), r.get("amount") or 0,
                 r.get("status") or "", r.get("source_id"), r.get("created_at") or now_iso()))
        conn.commit()
    except Exception as e:
        conn.rollback()
        conn.close()
        raise HTTPException(status_code=500, detail=f"Restore échoué : {e}")
    np_ = len(body["projects"]); no_ = len(body["operations"])
    conn.close()
    return {"ok": True, "projects": np_, "operations": no_}


# ---------------------------------------------------------------- projects

def _row_to_project(row: sqlite3.Row, today: Optional[date] = None, extra: Optional[dict] = None) -> dict:
    return project_computed(row, today, extra)


@app.get("/api/projects")
def list_projects(request: Request, platform: Optional[str] = None, status: Optional[str] = None):
    require_user(request)
    conn = db()
    q = "SELECT * FROM projects"
    conds, params = [], []
    if platform:
        conds.append("platform=?")
        params.append(platform)
    if status:
        conds.append("status=?")
        params.append(status)
    if conds:
        q += " WHERE " + " AND ".join(conds)
    q += " ORDER BY id DESC"
    rows = conn.execute(q, params).fetchall()
    # date du dernier revenu perçu (pour la fraîcheur des paiements)
    last_int = {}
    for r in conn.execute(
        "SELECT project_id, MAX(op_date) d FROM operations"
        " WHERE type LIKE '%Revenus%' AND amount > 0 AND status IN ('Validée','Réussi')"
        " GROUP BY project_id"):
        d = parse_date(r["d"])
        if d:
            last_int[r["project_id"]] = d
    # total reçu par projet (Σ montants positifs validés) + date de la dernière
    # opération — pour le taux réel annualisé des projets terminés
    recv_tot, last_op = {}, {}
    for r in conn.execute(
        "SELECT project_id, SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END) recv, MAX(op_date) d"
        " FROM operations WHERE status IN ('Validée','Réussi') AND project_id IS NOT NULL"
        " GROUP BY project_id"):
        recv_tot[r["project_id"]] = r["recv"] or 0
        last_op[r["project_id"]] = r["d"]
    conn.close()
    out = []
    for r in rows:
        p = _row_to_project(r, extra={
            "last_interest": last_int.get(r["id"]),
            "total_received": recv_tot,
            "last_op_date": last_op,
        })
        out.append(p)
    return {"projects": out}


@app.get("/api/projects/{pid}")
def get_project(pid: int, request: Request):
    require_user(request)
    conn = db()
    row = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    if row is None:
        raise HTTPException(status_code=404, detail="Projet introuvable")
    return _row_to_project(row)


@app.post("/api/projects")
def create_project(body: ProjectIn, request: Request):
    require_user(request)
    if body.platform not in PLATFORMS:
        raise HTTPException(status_code=400, detail="Plateforme inconnue")
    if body.status not in STATUS_LABELS:
        raise HTTPException(status_code=400, detail="Statut inconnu")
    conn = db()
    cur = conn.execute(
        """INSERT INTO projects
           (platform, name, city, invested, rate, duration_months, start_date,
            expected_end_date, actual_end_date, status, repaid_capital,
            interest_received, reinvested_from, notes, created_at, updated_at)
           VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?,?,?,?)""",
        (
            body.platform, body.name.strip(), body.city.strip(), body.invested,
            body.rate, body.duration_months, body.start_date, body.expected_end_date,
            body.actual_end_date, body.status, body.repaid_capital,
            body.interest_received, body.reinvested_from, body.notes.strip(),
            now_iso(), now_iso(),
        ),
    )
    conn.commit()
    pid = cur.lastrowid
    row = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    return _row_to_project(row)


@app.put("/api/projects/{pid}")
def update_project(pid: int, body: ProjectIn, request: Request):
    require_user(request)
    if body.platform not in PLATFORMS:
        raise HTTPException(status_code=400, detail="Plateforme inconnue")
    if body.status not in STATUS_LABELS:
        raise HTTPException(status_code=400, detail="Statut inconnu")
    conn = db()
    row = conn.execute("SELECT id FROM projects WHERE id=?", (pid,)).fetchone()
    if row is None:
        conn.close()
        raise HTTPException(status_code=404, detail="Projet introuvable")
    conn.execute(
        """UPDATE projects SET platform=?, name=?, city=?, invested=?, rate=?,
          duration_months=?, start_date=?, expected_end_date=?, actual_end_date=?,
          status=?, repaid_capital=?, interest_received=?, interest_net=?,
          interest_remaining=?, interest_remaining_net=?, real_rate=?,
          contract_type=?, valuation=?, reinvested_from=?, notes=?, updated_at=?
          WHERE id=?""",
        (
            body.platform, body.name.strip(), body.city.strip(), body.invested,
            body.rate, body.duration_months, body.start_date, body.expected_end_date,
            body.actual_end_date, body.status, body.repaid_capital,
            body.interest_received, body.interest_net, body.interest_remaining,
            body.interest_remaining_net, body.real_rate, body.contract_type,
            body.valuation, body.reinvested_from, body.notes.strip(), now_iso(), pid,
        ),
    )
    conn.commit()
    row = conn.execute("SELECT * FROM projects WHERE id=?", (pid,)).fetchone()
    conn.close()
    return _row_to_project(row)


@app.delete("/api/projects/{pid}")
def delete_project(pid: int, request: Request):
    require_user(request)
    conn = db()
    # neutraliser les liens de réinvestissement
    conn.execute("UPDATE projects SET reinvested_from=NULL WHERE reinvested_from=?", (pid,))
    cur = conn.execute("DELETE FROM projects WHERE id=?", (pid,))
    conn.commit()
    conn.close()
    if cur.rowcount == 0:
        raise HTTPException(status_code=404, detail="Projet introuvable")
    return {"ok": True}


# ---------------------------------------------------------------- summary

@app.get("/api/summary")
def summary(request: Request):
    require_user(request)
    conn = db()
    rows = conn.execute("SELECT * FROM projects").fetchall()
    # total déposé par plateforme (performance globale)
    metas = {}
    for r in conn.execute("SELECT platform, deposited FROM platform_meta"):
        metas[r["platform"]] = r["deposited"] or 0
    conn.close()
    today = date.today()

    total_invested = 0.0
    capital_due = 0.0
    interest_received = 0.0
    accrued = 0.0
    loss = 0.0
    unrealized = 0.0
    reinvested = 0.0
    n_total = len(rows)
    n_active = 0
    n_late = 0
    n_repaid = 0
    n_lost = 0
    by_platform: dict[str, dict] = {}

    for r in rows:
        c = project_computed(r, today)
        total_invested += c["invested"]
        capital_due += c["capital_due"]
        interest_received += c["interest_received"]
        accrued += c["accrued_interest"]
        loss += c["loss"]
        unrealized += c["unrealized_loss"]
        if c["status"] == "en_cours":
            n_active += 1
        if c["is_late"]:
            n_late += 1
        if c["status"] == "rembourse":
            n_repaid += 1
        if c["status"] == "perdu":
            n_lost += 1
        pkey = c["platform"] or "autre"
        bp = by_platform.setdefault(
            pkey, {"label": c["platform_label"], "invested": 0.0, "capital_due": 0.0,
                   "gain": 0.0, "loss": 0.0, "count": 0}
        )
        bp["invested"] += c["invested"]
        bp["capital_due"] += c["capital_due"]
        bp["gain"] += c["interest_received"] + c["accrued_interest"]
        bp["loss"] += c["loss"]
        bp["count"] += 1

    total_gain = round(interest_received + accrued, 2)
    net = round(total_gain - loss, 2)
    # « Capital recyclé » = investi au-delà des dépôts (remboursements + intérêts
    # réinvestis) — calculé par plateforme : max(0, investi − déposé)
    reinvested = round(sum(max(0.0, bp["invested"] - metas.get(pkey, 0)) for pkey, bp in by_platform.items()), 2)

    return {
        "total_invested": round(total_invested, 2),
        "capital_due": round(capital_due, 2),
        "interest_received": round(interest_received, 2),
        "accrued_interest": round(accrued, 2),
        "total_gain": total_gain,
        "loss": round(loss, 2),
        "unrealized_loss": round(unrealized, 2),
        "net": net,
        "reinvested": round(reinvested, 2),
        "counts": {
            "total": n_total, "active": n_active, "late": n_late,
            "repaid": n_repaid, "lost": n_lost,
        },
        "by_platform": [
            {"key": k, **v} for k, v in sorted(by_platform.items(), key=lambda kv: -kv[1]["invested"])
        ],
    }


# ---------------------------------------------------------------- export CSV

@app.get("/api/export.csv")
def export_csv(request: Request):
    require_user(request)
    conn = db()
    rows = conn.execute("SELECT * FROM projects ORDER BY start_date DESC, id DESC").fetchall()
    conn.close()

    buf = io.StringIO()
    buf.write("\ufeff")  # BOM pour Excel
    writer = csv.writer(buf, delimiter=";")
    writer.writerow([
        "Plateforme", "Projet", "Ville", "Mise (€)", "Taux brut (%)", "Taux réel annualisé (%)", "Durée (mois)",
        "Date investissement", "Échéance prévue", "Date fin réelle", "Statut",
        "Capital remboursé (€)", "Intérêts perçus (€)", "Intérêts latents (€)",
        "Perte (€)", "Gain total (€)", "Net (€)", "Retard (jours)", "Réinvesti de",
        "Notes",
    ])
    today = date.today()
    recv_tot, last_op = {}, {}
    for r in conn.execute(
        "SELECT project_id, SUM(CASE WHEN amount > 0 THEN amount ELSE 0 END) recv, MAX(op_date) d"
        " FROM operations WHERE status IN ('Validée','Réussi') AND project_id IS NOT NULL"
        " GROUP BY project_id"):
        recv_tot[r["project_id"]] = r["recv"] or 0
        last_op[r["project_id"]] = r["d"]
    for r in rows:
        c = project_computed(r, today, {
            "total_received": recv_tot,
            "last_op_date": last_op,
        })
        writer.writerow([
            c["platform_label"], c["name"], c["city"], f"{c['invested']:.2f}".replace(".", ","),
            f"{c['rate']:.2f}".replace(".", ","),
            (f"{(c['real_annual_pct'] or 0)*100:.2f}".replace(".", ",") if c["real_annual_pct"] is not None else ""),
            c["duration_months"],
            c["start_date"], c["expected_end_date"], c["actual_end_date"],
            STATUS_LABELS.get(c["status"], c["status"]),
            f"{c['repaid_capital']:.2f}".replace(".", ","),
            f"{c['interest_received']:.2f}".replace(".", ","),
            f"{c['accrued_interest']:.2f}".replace(".", ","),
            f"{c['loss']:.2f}".replace(".", ","),
            f"{c['total_gain']:.2f}".replace(".", ","),
            f"{c['net']:.2f}".replace(".", ","),
            c["days_delayed"],
            c["reinvested_from"] or "",
            c["notes"],
        ])
    data = buf.getvalue().encode("utf-8")
    return Response(
        content=data,
        media_type="text/csv; charset=utf-8",
        headers={"Content-Disposition": 'attachment; filename="crowdfunding-tracker.csv"'},
    )


# ---------------------------------------------------------------- import + operations

@app.post("/api/import")
async def import_xlsx(request: Request):
    """Importe un export xlsx (Bricks.co ou La Première Brique). Multipart : fichier 'file'."""
    require_user(request)
    form = await request.form()
    upload = form.get("file")
    if upload is None or not hasattr(upload, "filename"):
        raise HTTPException(status_code=400, detail="Fichier manquant")
    raw = await upload.read()
    if not raw:
        raise HTTPException(status_code=400, detail="Fichier vide")
    try:
        import openpyxl
        wb = openpyxl.load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
        ws = wb[wb.sheetnames[0]]
        rows = list(ws.iter_rows(values_only=True))
    except Exception:
        raise HTTPException(status_code=400, detail="Fichier xlsx illisible")
    if not rows:
        raise HTTPException(status_code=400, detail="Fichier vide")
    headers = rows[0]
    platform = _detect_platform(headers)
    if platform is None:
        raise HTTPException(status_code=400, detail="Format de fichier non reconnu (attendu : export Bricks.co ou La Première Brique)")
    data_rows = rows[1:]
    ops = _parse_bricks_rows(data_rows) if platform == "bricks" else _parse_lpb_rows(data_rows)
    conn = db()
    summary = _import_operations(conn, platform, ops)
    conn.close()
    return {"ok": True, "platform": platform, "platform_label": PLATFORMS[platform],
            "total_rows": len(data_rows), **summary}


@app.get("/api/operations")
def list_operations(request: Request, platform: Optional[str] = None,
                    project_id: Optional[int] = None, type: Optional[str] = None,
                    q: Optional[str] = None, limit: int = 200, offset: int = 0):
    require_user(request)
    limit = max(1, min(limit, 1000))
    conn = db()
    conds, params = [], []
    if platform:
        conds.append("o.platform=?")
        params.append(platform)
    if project_id:
        conds.append("o.project_id=?")
        params.append(project_id)
    if type:
        conds.append("o.type LIKE ?")
        params.append(f"%{type}%")
    if q:
        conds.append("(o.type LIKE ? OR o.details LIKE ? OR COALESCE(p.name,'') LIKE ?)")
        like = f"%{q}%"
        params += [like, like, like]
    where = (" WHERE " + " AND ".join(conds)) if conds else ""
    total = conn.execute(f"SELECT COUNT(*) c FROM operations o LEFT JOIN projects p ON p.id=o.project_id{where}", params).fetchone()["c"]
    rows = conn.execute(
        f"""SELECT o.*, p.name AS project_name FROM operations o
            LEFT JOIN projects p ON p.id=o.project_id{where}
            ORDER BY o.op_date DESC, o.id DESC LIMIT ? OFFSET ?""",
        params + [limit, offset],
    ).fetchall()
    conn.close()
    items = []
    for r in rows:
        d = dict(r)
        try:
            d["extra"] = __import__("json").loads(d["extra"] or "{}")
        except Exception:
            d["extra"] = {}
        d["platform_label"] = PLATFORMS.get(d["platform"], d["platform"])
        items.append(d)
    return {"total": total, "items": items, "limit": limit, "offset": offset}


@app.get("/api/operations/stats")
def operations_stats(request: Request):
    require_user(request)
    conn = db()
    rows = conn.execute(
        """SELECT o.platform, o.project_id, p.name AS project_name,
                  SUM(CASE WHEN o.amount >= 0 THEN o.amount ELSE 0 END) AS inc,
                  SUM(CASE WHEN o.amount < 0 THEN -o.amount ELSE 0 END) AS out,
                  COUNT(*) AS n
           FROM operations o LEFT JOIN projects p ON p.id=o.project_id
           WHERE o.status IN ('Validée', 'Réussi')
           GROUP BY o.platform, o.project_id"""
    ).fetchall()
    conn.close()
    by_platform: dict[str, dict] = {}
    by_project = []
    total_in = total_out = 0.0
    for r in rows:
        inc, out = r["inc"] or 0.0, r["out"] or 0.0
        total_in += inc
        total_out += out
        bp = by_platform.setdefault(r["platform"], {"label": PLATFORMS.get(r["platform"], r["platform"]), "in": 0.0, "out": 0.0, "n": 0})
        bp["in"] += inc
        bp["out"] += out
        bp["n"] += r["n"]
        if r["project_id"]:
            by_project.append({"project_id": r["project_id"], "project_name": r["project_name"],
                               "in": inc, "out": out, "n": r["n"]})
    by_project.sort(key=lambda x: -(x["out"] + x["in"]))
    return {"total_in": round(total_in, 2), "total_out": round(total_out, 2),
            "net": round(total_in - total_out, 2),
            "by_platform": list(by_platform.values()), "by_project": by_project}


# ---------------------------------------------------------------- scrape (extension navigateur)

@app.get("/api/sync/token")
def sync_token(request: Request):
    """Retourne (et crée si besoin) le token de synchronisation pour l'extension navigateur."""
    require_user(request)
    conn = db()
    row = conn.execute("SELECT token FROM sync_tokens WHERE id=1").fetchone()
    if row is None:
        token = secrets.token_urlsafe(32)
        conn.execute("INSERT INTO sync_tokens (id, token, created_at) VALUES (1, ?, ?)", (token, now_iso()))
        conn.commit()
        row = conn.execute("SELECT token FROM sync_tokens WHERE id=1").fetchone()
    conn.close()
    return {"token": row["token"]}


def _check_sync_token(request: Request) -> bool:
    tok = request.headers.get("X-Sync-Token", "")
    if not tok:
        return False
    conn = db()
    row = conn.execute("SELECT 1 FROM sync_tokens WHERE id=1 AND token=?", (tok,)).fetchone()
    conn.close()
    return row is not None


# clés candidates pour l'extraction heuristique
_RATE_KEYS = ("rate", "interest_rate", "annual_rate", "annual_interest", "taux", "gross_rate", "yield", "rendement", "rentability")
_DURATION_KEYS = ("duration", "duration_months", "months", "term", "duree", "horizon", "duration_in_months", "maturity_months")
_STATUS_KEYS = ("status", "state", "statut", "phase", "project_status", "current_status")
_INVESTED_KEYS = ("invested", "invested_amount", "total_invested", "amount_invested", "capital_invested", "mise", "total_invested_amount", "investment_amount", "montant_investi", "capital_investi", "montant_invest", "invested_capital")
_NAME_KEYS = ("name", "title", "project_name", "property_name", "label", "nom", "project", "property", "program_name", "slug", "property_name")
_AMOUNT_KEYS = ("amount", "total_amount", "amount_invested", "capital", "invested", "current_amount", "montant", "principal")
_DATE_KEYS = ("end_date", "expected_end_date", "maturity_date", "expected_date", "due_date", "fin_date", "term_date", "date_end")
_START_KEYS = ("start_date", "investment_date", "date_start", "begin_date", "purchase_date", "acquisition_date")


def _norm_scrape(s):
    s = unicodedata.normalize("NFD", str(s or ""))
    s = "".join(c for c in s if unicodedata.category(c) != "Mn")
    return s.lower().replace("’", "'").replace("‘", "'").strip()


def _to_float(v):
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).replace("\u00a0", " ").replace(" ", "").replace("\u20ac", "").replace("€", "").replace("%", "").replace(",", ".")
    s = s.strip()
    if not s:
        return None
    try:
        return float(s)
    except ValueError:
        return None


def _extract_project_fields(obj: dict) -> dict:
    """Extrait les champs projet d'un dict quelconque (heuristique par noms de clés)."""
    out = {}
    low = {str(k).lower().strip(): v for k, v in obj.items()}
    # nom
    for k in _NAME_KEYS:
        if k in low and isinstance(low[k], str) and low[k].strip():
            out["name"] = low[k].strip()
            break
    if "name" not in out and "id" in low and isinstance(low["id"], str) and low["id"].strip():
        out["name"] = low["id"].strip()  # slug utilisé en dernier recours
    # taux
    for k in _RATE_KEYS:
        if k in low:
            v = _to_float(low[k])
            if v is not None:
                out["rate"] = v
                break
    # durée
    for k in _DURATION_KEYS:
        if k in low:
            v = _to_float(low[k])
            if v is not None:
                out["duration_months"] = v
                break
    # statut
    for k in _STATUS_KEYS:
        if k in low and isinstance(low[k], str) and low[k].strip():
            out["status"] = low[k].strip().lower()
            break
    # investi / montant
    for k in _INVESTED_KEYS:
        if k in low:
            v = _to_float(low[k])
            if v is not None:
                out["invested"] = v
                break
    if "invested" not in out:
        for k in _AMOUNT_KEYS:
            if k in low:
                v = _to_float(low[k])
                if v is not None:
                    out["invested"] = v
                    break
    # dates
    for k in _DATE_KEYS:
        if k in low and isinstance(low[k], str) and low[k].strip():
            d = _parse_fr_date(low[k]) or _parse_iso_date(low[k])
            if d:
                out["expected_end_date"] = d
                break
    for k in _START_KEYS:
        if k in low and isinstance(low[k], str) and low[k].strip():
            d = _parse_fr_date(low[k]) or _parse_iso_date(low[k])
            if d:
                out["start_date"] = d
                break
    return out


def _parse_iso_date(s):
    import datetime as _dt
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return _dt.datetime.strptime(str(s).strip(), fmt).date().isoformat()
        except ValueError:
            continue
    return None


def _walk_json(node, out, parent: dict | None = None):
    """Parcourt récursivement une structure JSON et collecte les objets 'projet-like'.

    Si un objet projet-like (ex. `property`) est imbriqué dans un dict parent qui
    porte des champs financiers (ex. `invested_amount`), on fusionne parent+enfant
    pour ne pas perdre le montant (pattern API Bricks.co).
    """
    if isinstance(node, dict):
        cand = _extract_project_fields(node)
        # compléter avec les champs du parent (pattern Bricks : property + invested_amount au même niveau)
        if parent:
            for k in ("invested", "rate", "duration_months", "expected_end_date", "start_date", "status"):
                if k not in cand and k in parent:
                    cand[k] = parent[k]
        has_name = "name" in cand
        if not has_name and parent and "name" in parent:
            cand["name"] = parent["name"]
            has_name = True
        if has_name and ("rate" in cand or "duration_months" in cand or "invested" in cand or "status" in cand):
            out.append(cand)
        for k, v in node.items():
            if isinstance(v, dict):
                _walk_json(v, out, parent={**node, **cand})
            elif isinstance(v, list):
                _walk_json(v, out, parent)
    elif isinstance(node, list):
        for v in node:
            _walk_json(v, out, parent)


_SC_STATUS_MAP = {
    "en_cours": ("en_cours", "invested", "active", "funded", "financé", "finance", "running", "ongoing", "en cours", "en remboursement", "succeeded", "succes"),
    "en_collecte": ("en_collecte", "collecting", "open", "collecte", "en collecte", "fundraising", "en cours de collecte"),
    "rembourse": ("rembourse", "repaid", "completed", "closed", "finished", "terminé", "termine", "clos", "remboursé", "fini", "refunded", "rembourse"),
    "perdu": ("perdu", "default", "lost", "failure", "perte", "défaut", "defaut", "litige", "en procédure", "late"),
}


_MONTHS_FR = {
    "janvier": "01", "fevrier": "02", "mars": "03", "avril": "04", "mai": "05", "juin": "06",
    "juillet": "07", "aout": "08", "septembre": "09", "octobre": "10", "novembre": "11", "decembre": "12",
    "janv": "01", "fevr": "02", "mars": "03", "avr": "04", "mai": "05", "juin": "06",
    "juil": "07", "aout": "08", "sept": "09", "oct": "10", "nov": "11", "dec": "12",
    # abréviations 3 lettres (LPB : « 28 oct. 2024 », « 31 jan. 2024 », « 26 mar. 2026 »)
    "jan": "01", "fev": "02", "mar": "03", "jun": "06", "jui": "07", "aou": "08", "sep": "09",
}


def _extract_from_text(text: str, projects, global_ctx: bool = False) -> list:
    """Cherche chaque projet connu dans un texte brut (innerText d'une page détail)
    et extrait les infos disponibles dans son voisinage : taux %, durée mois, échéance.
    Si global_ctx=True (page détail d'UN projet), une passe globale sur tout le texte
    complète : revenus cumulés bruts/nets, revenus restants, taux réel, taux cible et
    échéance via « Remboursement final <Mois> <Année> » (souvent loin du nom)."""
    if not text:
        return []
    flat = re.sub(r"\s+", " ", text)
    hay = _norm_scrape(flat)
    out = []
    for p in projects:
        name = p["name"]
        if not name or len(name) < 3:
            continue
        # recherche du nom normalisé (sans accents, apostrophes unifiées)
        needle = _norm_scrape(name)
        if not needle:
            continue
        idx = hay.find(needle)
        if idx < 0:
            continue
        # fenêtre de contexte : du nom jusqu'au prochain projet connu (les listes
        # LPB/Bricks empilent les blocs → 900 chars débordent sur le voisin).
        # Limite gauche : après la fin du bloc précédent (« dont intérêts : X € »)
        # pour ne pas remonter dans le bloc d'avant (hay normalisé : « interets »).
        ctx_start = max(0, idx - 400)
        mi = hay.rfind("dont interets :", 0, idx)
        if mi >= 0:
            e = hay.find("€", mi)
            if e >= 0:
                ctx_start = max(ctx_start, e + 1)
        ctx_end = idx + len(needle) + 1500
        for p2 in projects:
            n2 = _norm_scrape(p2["name"])
            if not n2 or p2["name"] == name:
                continue
            j2 = hay.find(n2, idx + len(needle))
            if j2 >= 0 and j2 < ctx_end:
                ctx_end = j2
        ctx = flat[ctx_start: ctx_end]
        item = {"name": name}
        # taux : UNIQUEMENT avec label explicite (évite « 31,4 % » = rendement cumulé)
        m = re.search(
            r"(?:taux|rendement annuel|rendement brut|annuel|par an|par mois)\s*:?\s*([\d][\d.,]*)\s*%", ctx, re.I)
        if m:
            v = _to_float(m.group(1))
            if v is not None:
                item["rate"] = v
        # durée : UNIQUEMENT avec label « durée » ET pas « restante » (mois OU années).
        # « durée de vie du contrat » est un label explicite → fiable même si
        # « Durée restante X » traîne dans le contexte (pages projet Bricks)
        m = re.search(
            r"durée de vie du contrat[^0-9]{0,80}?(\d{1,3})\s*mois", ctx, re.I)
        if m:
            item["duration_months"] = int(m.group(1))
        else:
            m = re.search(
                r"(?:durée du contrat|durée totale|durée prévue|durée du projet|durée)\s*:?\s*(\d{1,3})\s*mois", ctx, re.I)
            if m and not re.search(r"restant|restante|il y a", ctx, re.I):
                item["duration_months"] = int(m.group(1))
            else:
                # années : tolère le texte explicatif entre le label et le nombre.
                # Label « durée de vie du contrat » explicite → fiable même si
                # « Durée restante X ans » traîne dans le contexte (pages Royalties)
                m2 = re.search(
                    r"durée de vie du contrat[^0-9]{0,250}?(\d{1,2})\s*ans?", ctx, re.I)
                if m2:
                    item["duration_months"] = int(m2.group(1)) * 12
        # badges juste après le nom : « Royalties » et « Retard de paiement ».
        # Fenêtre = jusqu'au premier « Cumulé » (max 120) : les badges du projet
        # VOISIN arrivent après son propre « Cumulé » → pas de faux positifs.
        # (hay est normalisé : « cumule » sans accent)
        tail = hay[idx + len(needle): idx + len(needle) + 120]
        cut = tail.find(" cumule")
        if cut >= 0:
            tail = tail[:cut]
        if re.search(r"royalt", tail, re.I):
            item["contract_type"] = "royalty"
        if re.search(r"retard de paiement", tail, re.I):
            item["status"] = "retard"
        # échéance : mot-clé + date FR / ISO, OU « Remboursement final X € <Mois> <Année> »
        m = re.search(
            r"(?:échéance|écheance|jusqu'au|jusqu au|remboursement prévu|remboursement prevu|maturité|maturite|date de fin|date de fin|termine le|terminé le)"
            r"\s*:?\s*(\d{1,2}\s+[a-zA-Zéû]+\.?\s+\d{4}|\d{1,2}/\d{1,2}/\d{2,4}|\d{4}-\d{2}-\d{2})",
            ctx, re.I)
        if m:
            iso = _parse_fr_text_date(m.group(1))
            if iso:
                item["expected_end_date"] = iso
        if not m:
            # « Remboursement final X € <Mois> <Année> » = fin du contrat UNIQUEMENT
            # pour les prêts classiques. Pour un ROYALTY, c'est une revente
            # PARTIELLE d'un lot : le contrat dure 10 ans (Horizon) — ne jamais
            # prendre cette date comme échéance (faux retard type Belfort).
            m2 = re.search(r"remboursement final\s*[\d\s.,]*€?\s*([a-zA-Zéû]+\.?)\s+(\d{4})", ctx, re.I)
            if m2 and not re.search(r"royalt", ctx, re.I):
                iso = _parse_fr_text_date("1 " + m2.group(1) + " " + m2.group(2))
                if iso:
                    # échéance = fin du mois annoncé
                    y, mo = int(iso[:4]), int(iso[5:7])
                    import calendar as _cal
                    item["expected_end_date"] = f"{y}-{mo:02d}-{_cal.monthrange(y, mo)[1]:02d}"
        # taux LPB : « 13,25% / an sur 18 mois » (pourcentage AVANT le label)
        if "rate" not in item:
            m = re.search(r"([\d][\d.,]*)\s*%\s*/\s*an", ctx, re.I)
            if m:
                v = _to_float(m.group(1))
                if v is not None:
                    item["rate"] = v
        # durée LPB : « sur 18 mois » (jamais « X mois max. restants »)
        if "duration_months" not in item:
            m = re.search(r"sur\s*(\d{1,3})\s*mois", ctx, re.I)
            if m:
                item["duration_months"] = int(m.group(1))
        # Royalties : durée = « Horizon X ans » (10 ans) — la « durée de vie du
        # contrat » (6 ans…) est un horizon initial révisable, pas l'échéance
        if item.get("contract_type") == "royalty" and "duration_months" not in item:
            m = re.search(r"horizon\s*:?\s*(\d{1,2})\s*ans?", ctx, re.I)
            if m:
                item["duration_months"] = int(m.group(1)) * 12
        # durée restante LPB : « 3 mois max. restants » — l'échéance OFFICIELLE
        # affichée par la plateforme (le compte à rebours démarre à la 1re
        # mensualité, PAS à la date de financement → « Financé + durée » est faux)
        m = re.search(r"(\d{1,3})\s*mois\s*max\.?\s*restants?", ctx, re.I)
        if m:
            item["rest_months"] = int(m.group(1))
        # paiement in-fine LPB : les intérêts sont versés au TERME du prêt
        # (pas de mensualités) → ne pas classer « critique » à tort
        if re.search(r"paiement in[- ]?fine", ctx, re.I):
            item["infine"] = 1
        # start_date LPB : « Financé le 28 oct. 2024 » (début du contrat)
        if "start_date" not in item:
            m = re.search(r"financé le\s*(\d{1,2})\s*([a-zéû]+\.?)\s*(\d{4})", ctx, re.I)
            if m:
                iso = _parse_fr_text_date(f"{m.group(1)} {m.group(2)} {m.group(3)}")
                if iso:
                    item["start_date"] = iso
        # intérêts LPB : « dont intérêts : 12,24 € » (capital hors intérêts)
        if "interest_received" not in item:
            m = re.search(r"dont intérêts\s*:?\s*([\d][\d\s.,]*)\s*€", ctx, re.I)
            if m:
                v = _to_float(m.group(1))
                if v is not None and v > 0:
                    item["interest_received"] = v
        # capital remboursé LPB : « Montant remboursé 101,24 € dont intérêts : 12,24 € »
        # → repaid_capital = 101,24 − 12,24 = 89,00 € (le capital, hors intérêts)
        if "repaid_capital" not in item:
            m = re.search(r"montant remboursé\s*([\d][\d\s.,]*)\s*€\s*dont intérêts\s*:?\s*([\d][\d\s.,]*)\s*€", ctx, re.I)
            if m:
                tot = _to_float(m.group(1))
                ints = _to_float(m.group(2))
                if tot is not None and ints is not None:
                    v = round(tot - ints, 2)
                    if v > 0:
                        item["repaid_capital"] = v
        # statut LPB : « Terminé le … » → remboursé
        if re.search(r"terminé le|termine le", ctx, re.I):
            item["status"] = "Terminé"
        # montant investi : label + montant
        m = re.search(r"(?:investi|montant investi|capital investi)\s*:?\s*([\d][\d\s.,]*)\s*€", ctx, re.I)
        if m:
            v = _to_float(m.group(1))
            if v is not None:
                item["invested"] = v
        # revenus cumulés (page Bricks) → intérêt reçu
        m = re.search(r"revenus cumulés\s*:?\s*[+]?([\d][\d\s.,]*)\s*€", ctx, re.I)
        if m:
            v = _to_float(m.group(1))
            if v is not None:
                item["interest_received"] = v
        out.append(item)

    # ---- passe globale (page détail d'UN projet) : les blocs sont loin du nom ----
    if global_ctx and out:
        g = out[0]
        # durée de vie du contrat (mois puis années) — page détail
        if "duration_months" not in g:
            m = re.search(r"durée de vie du contrat[^0-9]{0,250}?(\d{1,3})\s*mois", flat, re.I)
            if m:
                g["duration_months"] = int(m.group(1))
            else:
                m = re.search(r"durée de vie du contrat[^0-9]{0,250}?(\d{1,2})\s*ans?", flat, re.I)
                if m:
                    g["duration_months"] = int(m.group(1)) * 12
        # tags « Royalties » / « Retard de paiement » : après la DERNIÈRE occurrence
        # du nom (le détail de la page, pas le panneau latéral)
        if ("contract_type" not in g) or ("status" not in g):
            needle = _norm_scrape(g["name"])
            last = -1
            pos = 0
            while needle:
                j = hay.find(needle, pos)
                if j < 0:
                    break
                last = j
                pos = j + len(needle)
            if last >= 0:
                tail = hay[last + len(needle): last + len(needle) + 120]
                cut = tail.find(" cumule")
                if cut >= 0:
                    tail = tail[:cut]
                if "contract_type" not in g and re.search(r"royalt", tail, re.I):
                    g["contract_type"] = "royalty"
                if "status" not in g and re.search(r"retard de paiement", tail, re.I):
                    g["status"] = "retard"
        # Royalties : durée = « Horizon X ans » (10 ans), écrase la « durée de vie
        # du contrat » (horizon initial révisable — pas l'échéance)
        if g.get("contract_type") == "royalty":
            m = re.search(r"horizon\s*:?\s*(\d{1,2})\s*ans?", flat, re.I)
            if m:
                g["duration_months"] = int(m.group(1)) * 12
        # revenus cumulés brut + net (« +5,66 € / +3,97 € après fiscalité »)
        m = re.search(
            r"revenus cumulés\s*:?\s*[+]?([\d][\d\s.,]*)\s*€\s*[+]?([\d][\d\s.,]*)\s*€\s*après fiscalité",
            flat, re.I)
        if m:
            v1, v2 = _to_float(m.group(1)), _to_float(m.group(2))
            if v1 is not None:
                g["interest_received"] = v1
            if v2 is not None:
                g["interest_net"] = v2
        # revenus restants estimés (brut, puis net après fiscalité)
        m = re.search(
            r"revenus restants estimés\s*[^€]{0,120}?[+]?([\d][\d\s.,]*)\s*€(?:\s*[+]?([\d][\d\s.,]*)\s*€\s*après fiscalité)?",
            flat, re.I)
        if m:
            v1 = _to_float(m.group(1))
            if v1 is not None:
                g["interest_remaining"] = v1
            if m.group(2):
                v2 = _to_float(m.group(2))
                if v2 is not None:
                    g["interest_remaining_net"] = v2
        # NOTE : le « 31,4 % » sous « Revenus cumulés » est le TAUX D'IMPOSITION
        # (fiscalité), pas un rendement → volontairement PAS capturé comme real_rate.
        # taux cible / rentabilité annoncée
        m = re.search(
            r"(?:rentabilité cible|taux cible|rendement cible|taux annuel|taux brut)\s*:?\s*([\d][\d.,]*)\s*%",
            flat, re.I)
        if m:
            v = _to_float(m.group(1))
            if v is not None:
                g["rate"] = v
        # échéance : « Remboursement final 25,81 € Octobre 2026 » → fin de mois
        if "expected_end_date" not in g:
            m2 = re.search(r"remboursement final\s*[\d\s.,]*€?\s*([a-zA-Zéû]+\.?)\s+(\d{4})", flat, re.I)
            if m2:
                iso = _parse_fr_text_date("1 " + m2.group(1) + " " + m2.group(2))
                if iso:
                    import calendar as _cal
                    y, mo = int(iso[:4]), int(iso[5:7])
                    g["expected_end_date"] = f"{y}-{mo:02d}-{_cal.monthrange(y, mo)[1]:02d}"
        # montant investi de la page (label « Investi »/« Montant investi »)
        if "invested" not in g:
            m = re.search(r"(?:investi|montant investi|capital investi)\s*:?\s*([\d][\d\s.,]*)\s*€", flat, re.I)
            if m:
                v = _to_float(m.group(1))
                if v is not None:
                    g["invested"] = v
        # valeur actuelle des bricks (« Valeur de mes Bricks … 18,62 € »)
        m = re.search(r"valeur de mes bricks[^0-9]{0,160}?([\d][\d\s.,]*)\s*€", flat, re.I)
        if m:
            v = _to_float(m.group(1))
            if v is not None:
                g["valuation"] = v
    return out


def _parse_fr_text_date(s):
    """Parse une date française texte : '21 octobre 2024', '11 juil. 2025', '28 oct. 2024'."""
    import datetime as _dt
    s = str(s or "").strip()
    m = re.match(r"^(\d{1,2})\s+([a-zA-Zéû]+)\.?\s+(\d{4})$", s)
    if m:
        day, mon, year = m.groups()
        mon_key = unicodedata.normalize("NFD", mon.lower())
        mon_key = "".join(c for c in mon_key if unicodedata.category(c) != "Mn")
        if mon_key in _MONTHS_FR:
            try:
                return _dt.date(int(year), int(_MONTHS_FR[mon_key]), int(day)).isoformat()
            except ValueError:
                return None
    return _parse_iso_date(s) or _parse_fr_date(s)


def _parse_lpb_cards(cards) -> list:
    """Normalise les cards d'investissement LPB capturées par l'extension (DOM)."""
    out = []
    for c in cards or []:
        if not isinstance(c, dict):
            continue
        name = (c.get("name") or "").strip()
        if not name:
            continue
        item = {"name": name}
        inv = _to_float(c.get("invested"))
        if inv is not None:
            item["invested"] = inv
        rate = _to_float(c.get("rate"))
        if rate is not None:
            item["rate"] = rate
        dur = _to_float(c.get("duration_months") or c.get("duration"))
        if dur is not None:
            item["duration_months"] = dur
        status = (c.get("status") or "").strip()
        if status:
            item["status_raw"] = status
            item["status"] = status  # gardé pour rapport
        d = c.get("date") or c.get("start_date")
        if d:
            iso = _parse_fr_text_date(d)
            if iso:
                item["start_date"] = iso
        e = c.get("end_date")
        if e:
            iso = _parse_fr_text_date(e)
            if iso:
                item["expected_end_date"] = iso
        out.append(item)
    return out


def _map_status(s: str):
    s = s.strip().lower()
    if "retard de paiement" in s:
        return "retard"
    for canon, aliases in _SC_STATUS_MAP.items():
        for a in aliases:
            if a in s:
                return canon
    return None


@app.post("/api/scrape/ingest")
async def scrape_ingest(request: Request):
    """Reçoit les captures de l'extension (token X-Sync-Token), les stocke, tente
    d'enrichir les projets auto_created (taux/durée/échéance/statut) et produit un
    rapport de conformité site vs exports."""
    if not _check_sync_token(request):
        raise HTTPException(status_code=401, detail="Token de synchronisation invalide")
    try:
        data = await request.json()
    except Exception:
        raise HTTPException(status_code=400, detail="JSON invalide")
    captures = data.get("captures") or []
    if not isinstance(captures, list) or not captures:
        raise HTTPException(status_code=400, detail="Aucune capture reçue (captures: [])")

    conn = db()
    # stockage des captures brutes (limite 500)
    for c in captures[-500:]:
        try:
            body = json.dumps(c.get("body"), ensure_ascii=False)
        except Exception:
            body = str(c.get("body"))[:200000]
        conn.execute(
            "INSERT INTO scrape_captures (ts, platform, url, status_code, body) VALUES (?,?,?,?,?)",
            (c.get("ts") or now_iso(), (c.get("platform") or "")[:30], (c.get("url") or "")[:400],
             int(c.get("status") or 0), body),
        )

    # extraction heuristique des projets depuis toutes les captures
    found: dict[str, dict] = {}
    dom_captures = 0
    projects = conn.execute("SELECT * FROM projects").fetchall()

    def _absorb(items):
        for it in items:
            name = (it.get("name") or "").strip()
            if not name:
                continue
            key = _norm_scrape(name)
            if not key:
                continue
            if key not in found:
                found[key] = dict(it)
                continue
            # FUSION : l'item texte (taux, durée…) complète l'item cards (statut,
            # date) au lieu de le remplacer (le critère « len(it) > » jetait les
            # champs quand les deux items avaient autant de clés)
            cur = found[key]
            for k, v in it.items():
                if v is None or v == "" or v == 0:
                    continue  # ne jamais écraser avec une valeur vide/nulle
                if k in ("status", "status_raw", "name") and cur.get(k):
                    continue  # le statut des cards est plus riche que le texte
                cur[k] = v

    for c in captures:
        body = c.get("body")
        if not body:
            continue
        if isinstance(body, dict) and "_dom" in body:
            dom_captures += 1
            dom = body["_dom"]
            if isinstance(dom, dict):
                # cards structurées de la page (projet affiché)
                if dom.get("lpb_cards"):
                    _absorb(_parse_lpb_cards(dom["lpb_cards"]))
                if dom.get("cards"):
                    _absorb(_parse_lpb_cards(dom["cards"]))
                # texte brut : ne chercher QUE les projets affichés dans CETTE page
                # (les pages Bricks listent d'autres projets dans un panneau latéral
                #  → contexte croisé interdit : un projet ne peut être mis à jour
                #    que par sa propre page)
                if dom.get("inner_text"):
                    targets = []
                    for k in ("cards", "lpb_cards"):
                        for c in dom.get(k) or []:
                            if isinstance(c, dict) and c.get("name"):
                                targets.append(c["name"])
                    if not targets and dom.get("_bricks_project_page"):
                        # page projet Bricks sans cards (DOM expo, pas de h1/h2) :
                        # le projet affiché est celui dont le nom apparaît le plus
                        # (panneau latéral 1x + détail 1x = 2 occurrences minimum)
                        flat = dom["inner_text"].replace("\u00a0", " ")
                        norm_flat = _norm_scrape(flat)
                        counts = []
                        for p in projects:
                            needle = _norm_scrape(p["name"])
                            if needle and len(needle) >= 4:
                                n = norm_flat.count(needle)
                                if n > 0:
                                    counts.append((n, p["name"]))
                        if counts:
                            counts.sort(key=lambda x: -x[0])
                            top_n = counts[0][0]
                            # au moins 2 occurrences = nom du panneau + nom du détail
                            if top_n >= 2:
                                targets = [c[1] for c in counts if c[0] >= 2]
                    if targets:
                        tn = {_norm_scrape(t) for t in targets}
                        sub = [p for p in projects if _norm_scrape(p["name"]) in tn]
                        # page détail d'UN projet → passe globale (revenus nets,
                        # taux réel, échéance « Remboursement final »…)
                        _absorb(_extract_from_text(dom["inner_text"], sub,
                                                   global_ctx=len(sub) == 1))
                    else:
                        _absorb(_extract_from_text(dom["inner_text"], projects))
                # solde disponible de la plateforme (« Solde 519,85 € » sur les pages
                # LPB/Bricks) → stocké pour la performance globale
                flat_txt = re.sub(r"\s+", " ", dom.get("inner_text") or "")
                m = re.search(r"solde\s*:?\s*([\d][\d\s.,]*)\s*€", flat_txt, re.I)
                if m:
                    v = _to_float(m.group(1))
                    plat = (c.get("platform") or "").strip()
                    if v and v > 0 and plat in ("bricks", "lapremierebrique"):
                        row = conn.execute("SELECT balance FROM platform_meta WHERE platform=?", (plat,)).fetchone()
                        cur_bal = row["balance"] if row else 0
                        if abs((cur_bal or 0) - v) > 0.5:
                            conn.execute(
                                "INSERT INTO platform_meta (platform, balance, updated_at) VALUES (?,?,?)"
                                " ON CONFLICT(platform) DO UPDATE SET balance=?, updated_at=?",
                                (plat, v, now_iso(), v, now_iso()))
            continue
        items = []
        try:
            _walk_json(body, items)
        except Exception:
            continue
        _absorb(items)

    # matching + mise à jour des projets
    by_norm = {_norm_scrape(p["name"]): p for p in projects}
    updated, matched, new_fields = [], 0, 0
    for key, it in found.items():
        p = by_norm.get(key)
        if p is None:
            continue
        matched += 1
        updates, uparams = [], []
        if "rate" in it and (p["rate"] is None or p["rate"] <= 0):
            updates.append("rate=?")
            uparams.append(round(it["rate"], 2))
            new_fields += 1
        if "duration_months" in it and (p["duration_months"] is None or p["duration_months"] <= 0):
            updates.append("duration_months=?")
            uparams.append(int(it["duration_months"]))
            new_fields += 1
        # durée restante LPB (évolue chaque mois → mise à jour systématique)
        if "rest_months" in it and it["rest_months"] and it["rest_months"] != p["rest_months"]:
            updates.append("rest_months=?")
            uparams.append(int(it["rest_months"]))
            new_fields += 1
        # capital remboursé (évolue chaque mois → mise à jour systématique)
        if "repaid_capital" in it and it["repaid_capital"] != (p["repaid_capital"] or 0):
            updates.append("repaid_capital=?")
            uparams.append(round(it["repaid_capital"], 2))
            new_fields += 1
        if "interest_received" in it and (p["interest_received"] is None or p["interest_received"] <= 0):
            updates.append("interest_received=?")
            uparams.append(round(it["interest_received"], 2))
            new_fields += 1
        for extra_col in ("interest_net", "interest_remaining", "interest_remaining_net", "real_rate"):
            if extra_col in it and (p[extra_col] is None or p[extra_col] <= 0):
                updates.append(f"{extra_col}=?")
                uparams.append(round(it[extra_col], 2))
                new_fields += 1
        if "contract_type" in it and it["contract_type"] and not p["contract_type"]:
            updates.append("contract_type=?")
            uparams.append(it["contract_type"])
            new_fields += 1
        if "valuation" in it and it["valuation"] and (p["valuation"] is None or p["valuation"] <= 0):
            updates.append("valuation=?")
            uparams.append(round(it["valuation"], 2))
            new_fields += 1
        if "infine" in it and it["infine"] and not p["infine"]:
            updates.append("infine=1")
            new_fields += 1
        if "expected_end_date" in it and not p["expected_end_date"]:
            updates.append("expected_end_date=?")
            uparams.append(it["expected_end_date"])
            new_fields += 1
        if "start_date" in it and not p["start_date"]:
            updates.append("start_date=?")
            uparams.append(it["start_date"])
            new_fields += 1
        if "status" in it:
            mapped = _map_status(it["status"])
            if mapped and p["status"] in ("en_cours", "retard") and mapped != "en_cours" and mapped != p["status"]:
                updates.append("status=?")
                uparams.append(mapped)
        if updates:
            uparams.append(p["id"])
            conn.execute(f"UPDATE projects SET {', '.join(updates)}, updated_at=? WHERE id=?", uparams[:-1] + [now_iso(), p["id"]])
            updated.append(p["name"])

    # rapport de conformité : montant site vs montant exporté (opérations)
    report = []
    for key, it in found.items():
        p = by_norm.get(key)
        if p is None:
            continue
        site_inv = it.get("invested")
        if site_inv is None:
            continue
        # total investi réel selon nos opérations
        row = conn.execute(
            """SELECT SUM(-amount) t FROM operations
               WHERE project_id=? AND status IN ('Validée','Réussi') AND amount < 0""",
            (p["id"],),
        ).fetchone()
        export_inv = row["t"] or 0.0
        diff = round(site_inv - export_inv, 2)
        report.append({
            "project": p["name"],
            "platform": p["platform"],
            "site_invested": round(site_inv, 2),
            "export_invested": round(export_inv, 2),
            "diff": diff,
            "ok": abs(diff) < 0.01,
            "site_rate": it.get("rate"),
            "site_status": it.get("status"),
        })

    conn.commit()
    # mémorise le dernier rapport pour l'UI
    summ = {
        "captures": len(captures),
        "dom_captures": dom_captures,
        "projects_detected": len(found),
        "projects_matched": matched,
        "projects_updated": len(updated),
        "fields_filled": new_fields,
        "updated_names": updated[:50],
    }
    conn.execute(
        "INSERT OR REPLACE INTO scrape_report (id, data, created_at) VALUES (1, ?, ?)",
        (json.dumps({"summary": summ, "conformity": report}, ensure_ascii=False), now_iso()),
    )
    conn.commit()
    conn.close()
    return {
        "ok": True,
        "summary": summ,
        "conformity": sorted(report, key=lambda r: abs(r["diff"]), reverse=True)[:100],
    }


@app.get("/api/scrape/report")
def scrape_report(request: Request):
    """Dernier rapport de synchronisation (extension navigateur)."""
    require_user(request)
    conn = db()
    row = conn.execute("SELECT data, created_at FROM scrape_report WHERE id=1").fetchone()
    conn.close()
    if row is None:
        return {"ok": True, "report": None}
    try:
        data = json.loads(row["data"])
    except Exception:
        data = {}
    data["created_at"] = row["created_at"]
    return {"ok": True, "report": data}


@app.get("/api/version")
def version():
    vfile = BASE_DIR / "VERSION"
    ver = vfile.read_text().strip() if vfile.exists() else "0.0.0"
    return {"version": ver}


@app.get("/api/check-update")
def check_update(request: Request):
    """Compare la version locale à la dernière release GitHub (repo privé → token requis)."""
    require_user(request)
    repo = os.environ.get("GITHUB_REPO", "LostInTheBugs/crowdfunding-tracker")
    token = os.environ.get("GITHUB_TOKEN", "")
    vfile = BASE_DIR / "VERSION"
    current = vfile.read_text().strip() if vfile.exists() else "0.0.0"

    def key(v: str):
        parts = v.lstrip("v").split(".")[:3]
        while len(parts) < 3:
            parts.append("0")
        try:
            return tuple(int(x) for x in parts)
        except ValueError:
            return (0, 0, 0)

    try:
        headers = {"User-Agent": "crowdfunding-tracker"}
        if token:
            headers["Authorization"] = f"Bearer {token}"
        req = urllib.request.Request(
            f"https://api.github.com/repos/{repo}/releases/latest", headers=headers)
        with urllib.request.urlopen(req, timeout=8) as r:
            rel = json.load(r)
        latest = (rel.get("tag_name") or "").lstrip("v")
        return {
            "current": current,
            "latest": latest,
            "update_available": bool(latest) and key(latest) > key(current),
            "changelog": (rel.get("body") or "")[:1500],
            "url": rel.get("html_url") or "",
        }
    except Exception as e:
        return {"error": str(e)}


@app.get("/api/patrimoine-history")
def patrimoine_history(request: Request):
    """Série mensuelle du patrimoine (cash + encours au coût, calibrée sur le
    patrimoine actuel) depuis le 1er investissement, + mises cumulées."""
    require_user(request)
    conn = db()
    conn.row_factory = sqlite3.Row
    today = date.today()
    ops = conn.execute("""SELECT op_date, amount, project_id, type FROM operations
        WHERE status IN ('Validée','Réussi') ORDER BY op_date""").fetchall()
    metas = {r["platform"]: dict(r) for r in conn.execute("SELECT * FROM platform_meta")}
    cap_due = {}
    for r in conn.execute("SELECT platform, SUM(MAX(0, invested - COALESCE(repaid_capital,0))) s FROM projects GROUP BY platform"):
        cap_due[r["platform"]] = r["s"] or 0
    # ratio capital des remboursements par projet (capital vs intérêts)
    ratio_cap = {}
    for r in conn.execute("SELECT id, repaid_capital, interest_received FROM projects"):
        tot = (r["repaid_capital"] or 0) + (r["interest_received"] or 0)
        ratio_cap[r["id"]] = (r["repaid_capital"] or 0) / tot if tot > 0 else 1.0
    benchs = [dict(r) for r in conn.execute("SELECT * FROM benchmarks ORDER BY annual_pct DESC")]
    ih = {}
    for r in conn.execute("SELECT key, month, cum_bricks, cum_lpb FROM index_history"):
        ih.setdefault(r["key"], {})[r["month"]] = (r["cum_bricks"], r["cum_lpb"])
    conn.close()

    bal = {p: round((metas.get(p) or {}).get("balance") or 0, 2) for p in ("bricks", "lapremierebrique")}
    invval = round((metas.get("bricks") or {}).get("invested_value") or 0, 2)
    # patrimoine actuel (même définition que l'overview : bricks valeur saisie, LPB capital dû)
    pat_now = round(bal["bricks"] + invval + bal["lapremierebrique"] + cap_due.get("lapremierebrique", 0), 2)
    cash_now = round(bal["bricks"] + bal["lapremierebrique"], 2)

    events = []  # (date, delta_cash, delta_encours)
    for o in ops:
        try:
            d = date.fromisoformat(o["op_date"])
        except (TypeError, ValueError):
            continue
        amt = o["amount"] or 0
        delta_cash = amt  # toute op signée joue sur le solde
        delta_enc = 0.0
        if o["project_id"]:
            if amt < 0:  # mise / achat
                delta_enc = -amt
            elif o["type"] and "Revenus" in o["type"]:
                delta_enc = 0.0  # intérêts purs (Bricks)
            else:  # remboursement / revente : part capital selon le ratio du projet
                delta_enc = -amt * ratio_cap.get(o["project_id"], 1.0)
        events.append((d, delta_cash, delta_enc))

    if not events:
        return {"labels": [], "patrimoine": [], "mises": [], "current": pat_now}
    events.sort(key=lambda e: e[0])
    d0 = events[0][0]
    # grille mensuelle : fin de chaque mois de d0 à aujourd'hui
    months = []
    y, m = d0.year, d0.month
    while (y, m) <= (today.year, today.month):
        ny, nm = (y + 1, 1) if m == 12 else (y, m + 1)
        months.append(date(ny, nm, 1) - timedelta(days=1))
        y, m = ny, nm

    from bisect import bisect_right
    edates = [e[0] for e in events]
    tot_cash = sum(e[1] for e in events)
    labels, pats, mises = [], [], []
    cash_ts, enc_ts = [], []
    idx = 0
    run_cash = run_enc = run_mises = 0.0
    for t in months:
        i = bisect_right(edates, t)
        while idx < i:
            dc, de = events[idx][1], events[idx][2]
            run_cash += dc
            run_enc += de
            if dc < 0 and de > 0:  # mise / achat (les retraits n'en sont pas)
                run_mises += -dc
            idx += 1
        cash_t = round(cash_now - (tot_cash - run_cash), 2)
        cash_ts.append(cash_t)
        enc_ts.append(run_enc)
        labels.append(t.strftime("%Y-%m"))
        pats.append(round(cash_t + run_enc, 2))
        mises.append(round(run_mises, 2))
    # calibration finale : l'écart (PV latentes Bricks, répartition capital/intérêts
    # approximée…) est réparti au prorata de l'encours → la courbe finit sur pat_now
    enc_now = max(0.0, enc_ts[-1])
    ecart = pat_now - (cash_ts[-1] + enc_ts[-1])
    for i in range(len(pats)):
        ajust = ecart * (enc_ts[i] / enc_now) if enc_now > 0 else 0.0
        pats[i] = round(cash_ts[i] + enc_ts[i] + ajust, 2)
    # courbes simulées des indices (ETF Acc) : les dépôts réels placés à la date du
    # 1er achat de chaque plateforme — perfs cumulées réelles (index_history) ou
    # estimées depuis le taux annualisé (indices ajoutés, Livret A)
    dep_bricks = (metas.get("bricks") or {}).get("deposited") or 0
    dep_lpb = (metas.get("lapremierebrique") or {}).get("deposited") or 0
    d0b, d0l = date(2021, 9, 27), date(2022, 10, 3)
    indices = []
    for b in benchs:
        h = ih.get(b["key"], {})
        r = (b["annual_pct"] or 0) / 100
        vals = []
        for t in months:
            ms = t.strftime("%Y-%m")
            if ms in h:
                cb, cl = h[ms]
            else:
                cb = (1 + r) ** ((t - d0b).days / 365) - 1
                cl = (1 + r) ** ((t - d0l).days / 365) - 1
            # les dépôts ne courent qu'à partir de leur date de départ
            v_b = dep_bricks * (1 + cb) if t >= d0b else 0.0
            v_l = dep_lpb * (1 + cl) if t >= d0l else 0.0
            vals.append(round(v_b + v_l, 2))
        indices.append({"key": b["key"], "name": b["name"], "values": vals})
    return {"labels": labels, "patrimoine": pats, "mises": mises, "current": pat_now, "indices": indices}


@app.get("/api/platforms")
def platforms(request: Request):
    require_user(request)
    return {"platforms": PLATFORMS}


# ---------------------------------------------------------------- static

init_db()
_sync_stats = sync_indicators_from_ops(db())
app.mount("/", StaticFiles(directory=STATIC_DIR, html=True), name="static")

if __name__ == "__main__":
    import uvicorn

    uvicorn.run(app, host="0.0.0.0", port=PORT)
